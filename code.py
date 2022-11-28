from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
import openpyxl
from openpyxl import Workbook

def next_id(n):
    if n<10:
        num = '0' + str(n)
    else:
        num = str(n)
    return num

def next_page(driver, page_num):
    try:
        page_id_name = f"ctl00_ContentPlaceHolder1_ucSearchResults1_rptPaging_ctl{next_id(page_num)}_lnkButtonPaging"
        next_page_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, page_id_name)))
        next_page_tf = True
    except:
        next_page_button = None
        next_page_tf = False

    page_list = [next_page_tf, next_page_button]

    return page_list

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'jobsearch'
ws['A1'] = 'Job'
ws['B1'] = 'job_link'
r = 2

s = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=s)
driver.get("https://www.adecco.com.tw/advancedsearch.aspx?search=1")

category = True
n = 1 #從第n個類別開始抓（ex. n=1, 從第一個分類“MIS網管”開始依序向下抓取）

while category != None:
    try:
        category_id = f"ctl00_ContentPlaceHolderLeftNav_ucJobSearchFilter1_rptClassification_ctl{next_id(n)}_lbLink"
        category_name = driver.find_element(By.ID, f"ctl00_ContentPlaceHolderLeftNav_ucJobSearchFilter1_rptClassification_ctl{next_id(n)}_lbLink")\
            .get_attribute('textContent')
        category_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, category_id)))
        category_button.click()
        time.sleep(5)

        t = 1
        ws.cell(column=1, row=r, value=category_name)
        r += 1
        page_end = False
        while page_end == False:
            try:
                title = driver.find_element(By.XPATH, f'//*[@id="resultsList"]/div[{t}]/div[1]/a[1]') \
                    .get_attribute('textContent')
                link = driver.find_element(By.XPATH, f'//*[@id="resultsList"]/div[{t}]/div[1]/a[1]') \
                    .get_attribute('href')
                print(title+' | '+link)
                t += 1
                ws.cell(column=1, row=r, value=title)
                ws.cell(column=2, row=r, value=link)
                r += 1

            except:
                page_end = True

        p = 2
        while next_page(driver = driver, page_num = p)[0] == True:
            next_page(driver=driver, page_num=p)[1].click()
            page_end = False
            t = 1
            while page_end == False:
                try:
                    title = driver.find_element(By.XPATH, f'//*[@id="resultsList"]/div[{t}]/div[1]/a[1]')\
                        .get_attribute('textContent')
                    link = driver.find_element(By.XPATH, f'//*[@id="resultsList"]/div[{t}]/div[1]/a[1]') \
                        .get_attribute('href')
                    print(title+' | '+link)
                    t += 1
                    ws.cell(column=1, row=r, value=title)
                    ws.cell(column=2, row=r, value=link)
                    r += 1

                except:
                    page_end = True
            p += 1
        print(f'{category_name}  完成')

        driver.get("https://www.adecco.com.tw/advancedsearch.aspx?search=1")
        n += 1

    except:
        category = None
        driver.refresh()
        print('抓取結束')

wb.save(f'jobresults.xlsx')
